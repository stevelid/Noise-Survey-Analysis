#!/usr/bin/env python3
"""
Combine NTi single-measurement report files into parser-compatible *_Rpt_Report.txt
files for use in Noise Survey Analysis.

This is intended for folders containing sequences such as:
  *_123_Report.txt
  *_RTA_3rd_Report.txt

The dashboard parser already supports *_Rpt_Report.txt files, but deliberately
skips individual *_Report.txt files. This script bridges that gap for manual
measurement folders by producing combined overview files while leaving the
original files intact.
"""

from __future__ import annotations

import argparse
from dataclasses import dataclass
from pathlib import Path
from typing import Iterable
import re

from openpyxl import Workbook


TIME_START_PREFIX = "\tStart:"
TIME_END_PREFIX = "\tEnd:"


@dataclass
class BroadbandMeasurement:
    source_name: str
    start_date: str
    start_time: str
    end_date: str
    end_time: str
    values: list[str]


@dataclass
class RTAMeasurement:
    source_name: str
    start_date: str
    start_time: str
    end_date: str
    end_time: str
    metrics: list[tuple[str, list[str]]]


@dataclass
class OverviewBlock:
    title: str
    config_rows: list[tuple[str, str]]
    headers: list[list[str]]
    records: list[list[str]]
    keys: list[tuple[str, str]]


def _split_cells(line: str) -> list[str]:
    return line.rstrip("\n").split("\t")


def _find_line_index(lines: list[str], marker: str) -> int:
    for idx, line in enumerate(lines):
        if line.strip().startswith(marker):
            return idx
    raise ValueError(f"Could not find marker '{marker}'")


def _extract_time_value(lines: list[str], prefix: str) -> str:
    for line in lines:
        if line.strip().startswith(prefix.strip()):
            parts = _split_cells(line)
            non_empty = [part.strip() for part in parts if part.strip()]
            if len(non_empty) < 2:
                raise ValueError(f"Malformed time line: {line!r}")
            return non_empty[-1]
    raise ValueError(f"Could not find time field '{prefix.strip()}'")


def _parse_date_time(value: str) -> tuple[str, str]:
    match = re.match(r"^\s*(\d{4}-\d{2}-\d{2})\s*,\s*(\d{2}:\d{2}:\d{2})\s*$", value)
    if not match:
        raise ValueError(f"Could not parse date/time value: {value!r}")
    return match.group(1), match.group(2)


def _load_text(path: Path) -> list[str]:
    return path.read_text(encoding="utf-8", errors="ignore").splitlines()


def _extract_section(lines: list[str], marker: str) -> list[str]:
    start = _find_line_index(lines, marker) + 1
    output: list[str] = []
    for line in lines[start:]:
        stripped = line.strip()
        if stripped.startswith("#") and stripped != "#":
            break
        if stripped.startswith("#CheckSum"):
            break
        if stripped:
            output.append(line)
    return output


def _extract_config_rows(lines: list[str]) -> list[tuple[str, str]]:
    rows = _extract_section(lines, "# Hardware Configuration")
    output: list[tuple[str, str]] = []
    for row in rows:
        cells = [cell.strip() for cell in _split_cells(row)]
        non_empty = [cell for cell in cells if cell]
        if len(non_empty) >= 2:
            output.append((non_empty[0], non_empty[1]))
    return output


def _format_cell(value: str) -> str:
    cleaned = value.strip()
    if cleaned == "-.-":
        return "0"
    return cleaned


def _measurement_key(date_value: str, time_value: str) -> tuple[str, str]:
    return date_value.strip(), time_value.strip()


def _sheet_name_from_title(title: str) -> str:
    return f"NTI_{title}_Overview"


def _infer_job_number(path: Path) -> str | None:
    for parent in [path, *path.parents]:
        match = re.match(r"^(\d{3,})\b", parent.name)
        if match:
            return match.group(1)
    return None


def _build_overview_output_path(folder: Path) -> Path:
    job_number = _infer_job_number(folder)
    ref = folder.name
    if job_number:
        return folder / f"{job_number}.Overview.{ref}.xlsx"
    return folder / f"{ref}.Overview.xlsx"


def _write_block_header(ws, start_col: int, block: OverviewBlock) -> None:
    ws.cell(row=1, column=start_col, value=block.title)
    ws.cell(row=5, column=start_col, value="Hardware Configuration")
    for idx, (label, value) in enumerate(block.config_rows, start=6):
        ws.cell(row=idx, column=start_col + 1, value=label)
        ws.cell(row=idx, column=start_col + 2, value=value)
    for row_offset, header_row in enumerate(block.headers, start=12):
        for col_offset, value in enumerate(header_row):
            if value:
                ws.cell(row=row_offset, column=start_col + col_offset, value=value)


def _write_block_records(
    ws,
    start_col: int,
    block: OverviewBlock,
    ordered_keys: list[tuple[str, str]],
) -> None:
    record_lookup = {key: row for key, row in zip(block.keys, block.records)}
    for row_idx, key in enumerate(ordered_keys, start=15):
        row = record_lookup.get(key)
        if not row:
            continue
        for col_offset, value in enumerate(row):
            if value != "":
                ws.cell(row=row_idx, column=start_col + col_offset, value=value)


def _auto_size_columns(ws) -> None:
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            if cell.value is None:
                continue
            max_length = max(max_length, len(str(cell.value)))
        if max_length:
            ws.column_dimensions[column_letter].width = min(max_length + 2, 24)


def _write_overview_sheet(
    wb: Workbook,
    title: str,
    left_block: OverviewBlock | None,
    right_block: OverviewBlock | None,
) -> None:
    if left_block is None and right_block is None:
        return

    ws = wb.create_sheet(_sheet_name_from_title(title))

    left_width = len(left_block.headers[0]) if left_block else 0
    right_start = left_width + 3 if left_block else 1

    ordered_keys: list[tuple[str, str]] = []
    seen: set[tuple[str, str]] = set()
    for block in (left_block, right_block):
        if block is None:
            continue
        for key in block.keys:
            if key not in seen:
                seen.add(key)
                ordered_keys.append(key)
    ordered_keys.sort()

    if left_block:
        _write_block_header(ws, 1, left_block)
        _write_block_records(ws, 1, left_block, ordered_keys)
    if right_block:
        _write_block_header(ws, right_start, right_block)
        _write_block_records(ws, right_start, right_block, ordered_keys)

    _auto_size_columns(ws)


def _extract_common_prefix(lines: list[str], results_marker: str) -> list[str]:
    idx = _find_line_index(lines, results_marker)
    # Drop the # Time section from combined outputs; *_Rpt_Report files store
    # Start/Stop per row instead.
    prefix = lines[:idx]
    cleaned: list[str] = []
    skip_time = False
    for line in prefix:
        stripped = line.strip()
        if stripped == "# Time":
            skip_time = True
            continue
        if skip_time:
            if stripped.startswith("# ") and stripped != "# Time":
                skip_time = False
            else:
                continue
        if not skip_time:
            cleaned.append(line)
    return cleaned


def parse_single_broadband_report(path: Path) -> tuple[list[str], list[str], list[str], BroadbandMeasurement]:
    lines = _load_text(path)
    common_prefix = _extract_common_prefix(lines, "# Broadband Results")

    results_idx = _find_line_index(lines, "# Broadband Results")
    table_lines = [line for line in lines[results_idx + 1:] if line.strip() and not line.startswith("#CheckSum")]
    if len(table_lines) < 4:
        raise ValueError(f"Unexpected broadband table format in {path.name}")

    header_rows = table_lines[:3]
    data_row = _split_cells(table_lines[3])
    if len(data_row) < 5:
        raise ValueError(f"Broadband data row too short in {path.name}")
    start_date = data_row[1].strip()
    start_time = data_row[2].strip()
    end_date = data_row[3].strip()
    end_time = data_row[4].strip()

    measurement = BroadbandMeasurement(
        source_name=path.name,
        start_date=start_date,
        start_time=start_time,
        end_date=end_date,
        end_time=end_time,
        values=data_row[5:],
    )
    return common_prefix, header_rows, data_row[1:5], measurement


def parse_broadband_log(path: Path) -> tuple[list[tuple[str, str]], list[str], list[str], list[list[str]]]:
    lines = _load_text(path)
    config_rows = _extract_config_rows(lines)
    table_lines = _extract_section(lines, "# Broadband LOG Results")
    if len(table_lines) < 3:
        raise ValueError(f"Unexpected broadband log format in {path.name}")
    header_names = _split_cells(table_lines[0])[3:]
    header_units = _split_cells(table_lines[1])[3:]
    rows = [_split_cells(line) for line in table_lines[2:]]
    return config_rows, header_names, header_units, rows


def parse_rta_log(path: Path) -> tuple[list[tuple[str, str]], list[str], list[str], list[str], list[list[str]]]:
    lines = _load_text(path)
    config_rows = _extract_config_rows(lines)
    marker = "# RTA LOG Results L*eq_dt" if any(
        line.strip().startswith("# RTA LOG Results L*eq_dt") for line in lines
    ) else "# RTA LOG Results"
    table_lines = _extract_section(lines, marker)
    if len(table_lines) < 4:
        raise ValueError(f"Unexpected RTA log format in {path.name}")
    header_top = _split_cells(table_lines[0])[4:]
    header_mid = _split_cells(table_lines[1])[4:]
    header_units = _split_cells(table_lines[2])[4:]
    rows = [_split_cells(line) for line in table_lines[3:]]
    return config_rows, header_top, header_mid, header_units, rows


def parse_single_rta_report(path: Path) -> tuple[list[str], list[str], list[str], RTAMeasurement]:
    lines = _load_text(path)
    common_prefix = _extract_common_prefix(lines, "# RTA Results")

    time_start = _extract_time_value(lines, TIME_START_PREFIX)
    time_end = _extract_time_value(lines, TIME_END_PREFIX)
    start_date, start_time = _parse_date_time(time_start)
    end_date, end_time = _parse_date_time(time_end)

    results_idx = _find_line_index(lines, "# RTA Results")
    table_lines = [line for line in lines[results_idx + 1:] if line.strip() and not line.startswith("#CheckSum")]
    if len(table_lines) < 4:
        raise ValueError(f"Unexpected RTA table format in {path.name}")

    band_row = _split_cells(table_lines[0])
    unit_row = _split_cells(table_lines[1])
    metric_rows = [_split_cells(line) for line in table_lines[2:]]

    metrics: list[tuple[str, list[str]]] = []
    for row in metric_rows:
        if len(row) < 3:
            continue
        metric_name = row[1].strip()
        if not metric_name:
            continue
        metrics.append((metric_name, row[2:]))

    if not metrics:
        raise ValueError(f"No RTA metric rows found in {path.name}")

    measurement = RTAMeasurement(
        source_name=path.name,
        start_date=start_date,
        start_time=start_time,
        end_date=end_date,
        end_time=end_time,
        metrics=metrics,
    )
    return common_prefix, band_row, unit_row, measurement


def parse_repeat_broadband_report(path: Path) -> tuple[list[tuple[str, str]], list[str], list[str], list[list[str]]]:
    lines = _load_text(path)
    config_rows = _extract_config_rows(lines)
    table_lines = _extract_section(lines, "# Broadband Results")
    if len(table_lines) < 4:
        raise ValueError(f"Unexpected broadband rpt format in {path.name}")
    header_names = _split_cells(table_lines[1])[4:]
    header_units = _split_cells(table_lines[2])[4:]
    rows = [_split_cells(line) for line in table_lines[3:]]
    return config_rows, header_names, header_units, rows


def parse_repeat_rta_report(path: Path) -> tuple[list[tuple[str, str]], list[str], list[str], list[str], list[list[str]]]:
    lines = _load_text(path)
    config_rows = _extract_config_rows(lines)
    table_lines = _extract_section(lines, "# RTA Results")
    if len(table_lines) < 4:
        raise ValueError(f"Unexpected RTA rpt format in {path.name}")
    header_top = _split_cells(table_lines[0])[5:]
    header_mid = _split_cells(table_lines[1])[5:]
    header_units = _split_cells(table_lines[2])[5:]
    rows = [_split_cells(line) for line in table_lines[3:]]
    return config_rows, header_top, header_mid, header_units, rows


def build_broadband_log_block(paths: Iterable[Path]) -> OverviewBlock | None:
    files = sorted(paths)
    if not files:
        return None

    config_rows, header_names, header_units, _ = parse_broadband_log(files[0])
    headers = [
        ["File", "Start", "", ""] + [""] * len(header_names),
        ["", "Date", "Time", "Timer", *header_names],
        ["", "[YYYY-MM-DD]", "[hh:mm:ss]", "[hh:mm:ss]", *header_units],
    ]

    records: list[list[str]] = []
    keys: list[tuple[str, str]] = []
    for path in files:
        _, _, _, rows = parse_broadband_log(path)
        for idx, row in enumerate(rows):
            date_value = row[1].strip()
            time_value = row[2].strip()
            records.append(
                [path.name if idx == 0 else "", date_value, time_value, row[3].strip(), *[_format_cell(v) for v in row[4:]]]
            )
            keys.append(_measurement_key(date_value, time_value))

    return OverviewBlock("123_Log", config_rows, headers, records, keys)


def build_rta_log_block(paths: Iterable[Path]) -> OverviewBlock | None:
    files = sorted(paths)
    if not files:
        return None

    config_rows, header_top, header_mid, header_units, _ = parse_rta_log(files[0])
    headers = [
        ["File", "Start", "", ""] + header_top,
        ["", "Date", "Time", "Timer", *header_mid],
        ["", "[YYYY-MM-DD]", "[hh:mm:ss]", "[hh:mm:ss]", *header_units],
    ]

    records: list[list[str]] = []
    keys: list[tuple[str, str]] = []
    for path in files:
        _, _, _, _, rows = parse_rta_log(path)
        for idx, row in enumerate(rows):
            date_value = row[1].strip()
            time_value = row[2].strip()
            records.append(
                [path.name if idx == 0 else "", date_value, time_value, row[3].strip(), *[_format_cell(v) for v in row[5:]]]
            )
            keys.append(_measurement_key(date_value, time_value))

    first_name = files[0].name
    suffix = "RTA_Log"
    if "_RTA_" in first_name:
        suffix = first_name.split("_RTA_", 1)[1].replace(".txt", "")
        suffix = f"RTA_{suffix}"
    return OverviewBlock(suffix, config_rows, headers, records, keys)


def build_broadband_report_block(paths: Iterable[Path], title: str) -> OverviewBlock | None:
    files = sorted(paths)
    if not files:
        return None

    config_rows, header_rows, _, first_measurement = parse_single_broadband_report(files[0])
    header_names = _split_cells(header_rows[1])[4:]
    header_units = _split_cells(header_rows[2])[4:]
    headers = [
        ["File", "Start", "", "Stop", ""] + [""] * len(header_names),
        ["", "Date", "Time", "Date", "Time", *header_names],
        ["", "[YYYY-MM-DD]", "[hh:mm:ss]", "[YYYY-MM-DD]", "[hh:mm:ss]", *header_units],
    ]

    records = [[
        files[0].name,
        first_measurement.start_date,
        first_measurement.start_time,
        first_measurement.end_date,
        first_measurement.end_time,
        *[_format_cell(v) for v in first_measurement.values],
    ]]
    keys = [_measurement_key(first_measurement.start_date, first_measurement.start_time)]

    for path in files[1:]:
        _, _, _, measurement = parse_single_broadband_report(path)
        records.append([
            path.name,
            measurement.start_date,
            measurement.start_time,
            measurement.end_date,
            measurement.end_time,
            *[_format_cell(v) for v in measurement.values],
        ])
        keys.append(_measurement_key(measurement.start_date, measurement.start_time))

    return OverviewBlock(title, _extract_config_rows(_load_text(files[0])), headers, records, keys)


def build_rta_report_block(paths: Iterable[Path], title: str) -> OverviewBlock | None:
    files = sorted(paths)
    if not files:
        return None

    _, band_row, unit_row, first_measurement = parse_single_rta_report(files[0])
    config_rows = _extract_config_rows(_load_text(files[0]))
    freq_headers = band_row[2:]
    freq_units = unit_row[2:]
    metric_names = [name for name, _ in first_measurement.metrics]

    header_top = ["File", "Start", "", "Stop", ""]
    header_mid = ["", "Date", "Time", "Date", "Time"]
    header_units = ["", "[YYYY-MM-DD]", "[hh:mm:ss]", "[YYYY-MM-DD]", "[hh:mm:ss]"]
    for metric_name in metric_names:
        header_top.extend([metric_name] * len(freq_headers))
        header_mid.extend(freq_headers)
        header_units.extend(freq_units)

    def flatten_measurement(path_name: str, measurement: RTAMeasurement) -> list[str]:
        values: list[str] = [path_name, measurement.start_date, measurement.start_time, measurement.end_date, measurement.end_time]
        for _, metric_values in measurement.metrics:
            values.extend(_format_cell(v) for v in metric_values)
        return values

    records = [flatten_measurement(files[0].name, first_measurement)]
    keys = [_measurement_key(first_measurement.start_date, first_measurement.start_time)]

    for path in files[1:]:
        _, _, _, measurement = parse_single_rta_report(path)
        records.append(flatten_measurement(path.name, measurement))
        keys.append(_measurement_key(measurement.start_date, measurement.start_time))

    return OverviewBlock(title, config_rows, [header_top, header_mid, header_units], records, keys)


def build_broadband_repeat_block(path: Path) -> OverviewBlock:
    config_rows, header_names, header_units, rows = parse_repeat_broadband_report(path)
    headers = [
        ["File", "Start", "", "Stop", ""] + [""] * len(header_names),
        ["", "Date", "Time", "Date", "Time", *header_names],
        ["", "[YYYY-MM-DD]", "[hh:mm:ss]", "[YYYY-MM-DD]", "[hh:mm:ss]", *header_units],
    ]

    records: list[list[str]] = []
    keys: list[tuple[str, str]] = []
    for idx, row in enumerate(rows):
        start_date = row[1].strip()
        start_time = row[2].strip()
        records.append([
            path.name if idx == 0 else "",
            start_date,
            start_time,
            row[3].strip(),
            row[4].strip(),
            *[_format_cell(v) for v in row[5:]],
        ])
        keys.append(_measurement_key(start_date, start_time))

    return OverviewBlock("123_Rpt_Report", config_rows, headers, records, keys)


def build_rta_repeat_block(path: Path) -> OverviewBlock:
    config_rows, header_top, header_mid, header_units, rows = parse_repeat_rta_report(path)
    headers = [
        ["File", "Start", "", "Stop", ""] + header_top,
        ["", "Date", "Time", "Date", "Time", *header_mid],
        ["", "[YYYY-MM-DD]", "[hh:mm:ss]", "[YYYY-MM-DD]", "[hh:mm:ss]", *header_units],
    ]

    records: list[list[str]] = []
    keys: list[tuple[str, str]] = []
    for idx, row in enumerate(rows):
        start_date = row[1].strip()
        start_time = row[2].strip()
        records.append([
            path.name if idx == 0 else "",
            start_date,
            start_time,
            row[3].strip(),
            row[4].strip(),
            *[_format_cell(v) for v in row[6:]],
        ])
        keys.append(_measurement_key(start_date, start_time))

    first_name = path.name
    suffix = "RTA_Rpt_Report"
    if "_RTA_" in first_name:
        suffix = first_name.split("_RTA_", 1)[1].replace(".txt", "")
        suffix = f"RTA_{suffix}"
    return OverviewBlock(suffix, config_rows, headers, records, keys)


def build_overview_workbook(
    folder: Path,
    broadband_files: list[Path],
    rta_files: list[Path],
    broadband_output: Path | None,
    rta_output: Path | None,
) -> Path:
    wb = Workbook()
    wb.remove(wb.active)

    _write_overview_sheet(
        wb,
        "Log",
        build_broadband_log_block(folder.glob("*_123_Log.txt")),
        build_rta_log_block(folder.glob("*_RTA_*_Log.txt")),
    )
    _write_overview_sheet(
        wb,
        "Report",
        build_broadband_report_block(broadband_files, "123_Report"),
        build_rta_report_block(rta_files, "RTA_Report"),
    )
    _write_overview_sheet(
        wb,
        "Rpt_Report",
        build_broadband_repeat_block(broadband_output) if broadband_output else None,
        build_rta_repeat_block(rta_output) if rta_output else None,
    )

    output_path = _build_overview_output_path(folder)
    wb.save(output_path)
    return output_path


def _build_output_name(first_input: Path, suffix_from: str, suffix_to: str) -> str:
    return first_input.name.replace(suffix_from, suffix_to)


def combine_broadband_reports(paths: Iterable[Path]) -> Path:
    files = sorted(paths)
    if not files:
        raise ValueError("No broadband report files provided")

    prefix, header_rows, _, first_measurement = parse_single_broadband_report(files[0])
    measurements = [first_measurement]

    expected_value_count = len(first_measurement.values)
    for path in files[1:]:
        _, other_header_rows, _, measurement = parse_single_broadband_report(path)
        if other_header_rows != header_rows:
            raise ValueError(f"Broadband headers do not match: {path.name}")
        if len(measurement.values) != expected_value_count:
            raise ValueError(f"Broadband value count mismatch: {path.name}")
        measurements.append(measurement)

    output_path = files[0].with_name(_build_output_name(files[0], "_123_Report.txt", "_123_Rpt_Report.txt"))
    title_line = f"XL2 Sound Level Meter Broadband Reporting:\t\t{output_path.parent.name}\\{output_path.name}"

    content: list[str] = [title_line, "------------------------------------------", ""]
    content.extend(prefix[3:])  # keep shared metadata, skip original title/underline/blank
    content.append("# Broadband Results")
    content.extend(header_rows)
    for measurement in measurements:
        row = [
            "",
            measurement.start_date,
            measurement.start_time,
            measurement.end_date,
            measurement.end_time,
            *measurement.values,
        ]
        content.append("\t".join(row))
    content.extend(["", "#CheckSum", "\tGENERATED_BY_COMBINE_NTI_REPORTS"])
    output_path.write_text("\n".join(content) + "\n", encoding="utf-8")
    return output_path


def combine_rta_reports(paths: Iterable[Path]) -> Path:
    files = sorted(paths)
    if not files:
        raise ValueError("No RTA report files provided")

    prefix, band_row, unit_row, first_measurement = parse_single_rta_report(files[0])
    measurements = [first_measurement]

    expected_metric_names = [name for name, _ in first_measurement.metrics]
    expected_value_count = len(first_measurement.metrics[0][1])

    for path in files[1:]:
        _, other_band_row, other_unit_row, measurement = parse_single_rta_report(path)
        if other_band_row != band_row or other_unit_row != unit_row:
            raise ValueError(f"RTA headers do not match: {path.name}")
        metric_names = [name for name, _ in measurement.metrics]
        if metric_names != expected_metric_names:
            raise ValueError(f"RTA metric rows do not match: {path.name}")
        if any(len(values) != expected_value_count for _, values in measurement.metrics):
            raise ValueError(f"RTA band count mismatch: {path.name}")
        measurements.append(measurement)

    output_path = files[0].with_name(_build_output_name(files[0], "_Report.txt", "_Rpt_Report.txt"))
    title_line = f"XL2 Sound Level Meter RTA Reporting:\t\t{output_path.parent.name}\\{output_path.name}"

    content: list[str] = [title_line, "------------------------------------", ""]
    content.extend(prefix[3:])  # keep shared metadata, skip original title/underline/blank
    content.append("# RTA Results ")

    metric_names = expected_metric_names
    freq_headers = band_row[2:]
    units = unit_row[2:]

    header_top = ["", "Start", "", "Stop", ""]
    header_mid = ["", "Date", "Time", "Date", "Time"]
    header_bottom = ["", "[YYYY-MM-DD]", "[hh:mm:ss]", "[YYYY-MM-DD]", "[hh:mm:ss]"]

    for metric_name in metric_names:
        header_top.extend([metric_name, *([metric_name] * (len(freq_headers) - 1)), ""])
        header_mid.extend(["Band [Hz]", *freq_headers[1:], ""])
        header_bottom.extend(["", *units[1:], ""])

    content.append("\t".join(header_top))
    content.append("\t".join(header_mid))
    content.append("\t".join(header_bottom))

    for measurement in measurements:
        row = ["", measurement.start_date, measurement.start_time, measurement.end_date, measurement.end_time]
        for _, values in measurement.metrics:
            row.extend(["", *values])
        content.append("\t".join(row))

    content.extend(["", "#CheckSum", "\tGENERATED_BY_COMBINE_NTI_REPORTS"])
    output_path.write_text("\n".join(content) + "\n", encoding="utf-8")
    return output_path


def main() -> int:
    parser = argparse.ArgumentParser(description="Combine NTi single-measurement report files into *_Rpt_Report.txt files.")
    parser.add_argument("folder", help="Folder containing NTi *_Report.txt files")
    args = parser.parse_args()

    folder = Path(args.folder)
    if not folder.is_dir():
        raise SystemExit(f"Folder not found: {folder}")

    broadband_files = sorted(folder.glob("*_123_Report.txt"))
    rta_files = sorted(path for path in folder.glob("*_RTA_*_Report.txt") if "_Rpt_Report" not in path.name)

    if not broadband_files and not rta_files:
        raise SystemExit("No NTi single-measurement report files found.")

    broadband_output: Path | None = None
    rta_output: Path | None = None

    if broadband_files:
        broadband_output = combine_broadband_reports(broadband_files)
        print(f"[OK] Broadband combined file: {broadband_output}")

    if rta_files:
        rta_output = combine_rta_reports(rta_files)
        print(f"[OK] RTA combined file: {rta_output}")

    overview_output = build_overview_workbook(folder, broadband_files, rta_files, broadband_output, rta_output)
    print(f"[OK] Overview workbook: {overview_output}")

    return 0


if __name__ == "__main__":
    raise SystemExit(main())
